# Single particle/cell event identification algorithm
# Writen by Tian Xiangwei in Apr,2022
import openpyxl
import numpy
import copy
from alive_progress import alive_bar
import tkinter
from tkinter import filedialog
import os

print("Loading the file, wait......")
file0 = openpyxl.Workbook("Temporary empty file.xlsx")
origin_sheet0 = file0.create_sheet("origin event")
event_sheet0 = file0.create_sheet("event")
file0.save("Temporary empty file.xlsx")  # 先创建再读取，否则无法写入
file = openpyxl.load_workbook("Temporary empty file.xlsx")
origin_sheet = file["origin event"]
event_sheet = file["event"]


def openreadtxt(file_name):
    data = []
    file = open(file_name, 'r')  # 打开文件
    file_data = file.readlines()  # 读取所有行
    for row in file_data:
        tmp_list = row.split("\t")  # 按‘\t’切分每行的数据
        tmp_list[-1] = tmp_list[-1].replace('\n', ',')  # 去掉换行符
        data.append(tmp_list)  # 将每行数据插入data中
    return data


'''对话框选择获取文件路径和文件名,用于后续读取和命名'''
root = tkinter.Tk()  # 打开选择文件夹对话框
root.withdraw()
folderpath = filedialog.askdirectory()  # 获得选择好的文件夹
filepath = filedialog.askopenfilename()  # 获得选择好的文件
file_name0 = filepath.split("/")[-1]  # 去除读取的文件路径
file_name = os.path.splitext(os.path.basename(file_name0))[0]  # 去后缀

data_row = openreadtxt(f"{filepath}")
data_row_list = list(data_row)
data_column_list = [[] for element in range(0, len(data_row_list[10]))]
del data_row_list[0:7]
for row in data_row_list:
    for element in range(0, len(row)):
        data_column_list[element].append(row[element])
data_column_list = [ele for ele in data_column_list if ele != []]

selected_index = [2,  # 第二列时间列，须保留，更改下面的。是excel的列号，不是列表元素编号
                  4, 5, 7, 17, 18, 22, 24, 30, 37, 39, 46, 54, 55, 57, 62, 69, 71, 78, 79, 85, 86, 92, 93, 101, 102,
                  106, 113, 116, 126, 131, 132, 133, 137, 140, 150, 153, 156, 159, 165, 173, 177, 181, 189, 202, 209,
                  213, 214, 215, 222, 223, 232, 239, 242, 248, 251, 252, 256, 260, 264, 269, 274, 275, 279, 285,
                  291, 293, 296, 301, 302, 306, 309, 310, 312, 315]
selected_column_list = []
for i in selected_index:
    selected_column_list.append(data_column_list[i - 1])

row_numb = len(data_row_list)
col_numb = len(selected_column_list)
row_begin = 2  # 第二行第二列开始处理
col_begin = 2
window = 1000  # 确定计算窗口的数据点数目

print("Data processing......")
with alive_bar(col_numb - col_begin + 1, title="进度", force_tty=True) as bar:  # 进度条
    for element in range(col_begin - 1, col_numb):  # element为列序号，对不同元素所在的列逐列处理
        element_column_list = selected_column_list[element]

        bar()  # 进度条

        col_complete = (row_numb + window) * [None]  # 创建一个全为None的列表，后面识别出的事件强度填在相应的位置

        for i in range(row_begin - 1, row_numb, window):  # 每1000个数据点为一个处理单元，col列表填满这一千个数据点
            col = []
            for j in range(i, i + window):
                if j > (len(element_column_list) - 1):
                    if len(element_column_list) - i > 100:
                        break
                    else:
                        col = [None]
                        break
                cell = element_column_list[j]
                if cell != None:
                    cell = 65 * float(cell)  # 转换为count，若本就是count，不需要*65
                col.append(cell)

            if all(cell is None for cell in col):  # 行底端没有数据了，列表全空，意味着这一部分的计算结束了
                break

            cutoff = 100000
            upper_cutoff = cutoff + 1
            while cutoff != upper_cutoff:  # 1000的window内计算阈值
                col_filtered = list(filter(lambda x: x is not None, col))
                mean = numpy.average(col_filtered)
                std = numpy.std(col_filtered)
                upper_cutoff = cutoff
                cutoff = mean + 3.29 * std + 2.71;threhold_method = 'IUPAC'
                # cutoff = mean + 3 * std;threhold_method = 'μ+3σ'

                for k in range(0, len(col)):  # 迭代筛出事件，放到前面创建的全为None的列表中
                    if col[k] != None and col[k] > cutoff:
                        col_complete[i + k - row_begin] = col[k]
                        col[k] = None

            col_filtered = list(filter(lambda x: x is not None, col))  # 扣去背景值
            bgd = numpy.average(col_filtered)
            for k in range(0, window):
                if col_complete[i + k - 3] != None:
                    col_complete[i + k - 3] = col_complete[i + k - 3] - bgd

        for i in range(0, len(col_complete) - 1):  # 如果连续检出事件，将其加和成一个事件，当然背景还是在前面分别地减去
            if col_complete[i] != None and col_complete[i + 1] != None:
                col_complete[i] = col_complete[i] + col_complete[i + 1]
                col_complete[i + 1] = col_complete[i]  # 两格都填，后面会删除其中一格

        col_complete.insert(0, element_column_list[0])  # 横向表头-化学元素信息
        for i in range(0, len(col_complete)):  # 填进excel
            origin_sheet.cell(row=i + 1, column=element + 1).value = col_complete[i]
        for i in range(0, row_numb):
            time_header = selected_column_list[0]
            origin_sheet.cell(row=i + 1, column=1).value = time_header[i]

print("Summing the slipt events, wait......")
upper_row = col_numb * [None]
rows = []
event = []
for row in origin_sheet.values:
    rows.append(list(row))

for i in range(0, row_numb):  # 整合被分割的事件
    row = rows[i]
    event_count = len(event)
    index = True
    slipt_event = col_numb * [None]
    if all(cell is None for cell in row[1:col_numb]):
        continue
    for cell in range(1, col_numb):
        if row[cell] != None and row[cell] == upper_row[cell]:
            for cell in range(0, col_numb):
                if upper_row[cell] != None:
                    slipt_event[cell] = upper_row[cell]
                elif row[cell] != None:
                    slipt_event[cell] = row[cell]
            event[event_count - 1] = copy.deepcopy(slipt_event)
            index = False
            break
    upper_row = copy.deepcopy(row)
    if index == True:
        event.append(row)

for row in event:  # 符合要求的行，即含有事件的行，填入新的sheet中
    event_sheet.append(tuple(row))

print(f"Threhold method: {threhold_method}")
print("The file is saving, wait......")

del file["origin event"]  # 删除废sheet

file.save(f"{file_name}_{threhold_method}.xlsx")

print("\033[1mDone\n注意点：\n(1)适用于TOF的txt文件。\n(2)49行，选择要分析所在元素的列号，是表格列号不是python列表号\n"
      "(3)84行原始数据中强度转化*65，若本就是count，不用*65\n(4)97 98行，迭代算法的选择\033[0m")
