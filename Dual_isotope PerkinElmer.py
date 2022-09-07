
import openpyxl
import csv
import numpy
import tkinter
from tkinter import filedialog
import os
import copy

'''对话框选择获取文件路径和文件名,用于后续读取和命名'''
root = tkinter.Tk()  # 打开选择文件夹对话框
root.withdraw()
folderpath = filedialog.askdirectory()  # 获得选择好的文件夹
filepath = filedialog.askopenfilename()  # 获得选择好的文件
file_name0 = filepath.split("/")[-1]  # 去除读取的文件路径
file_name = os.path.splitext(os.path.basename(file_name0))[0]  # 去后缀
data_row_list = list(csv.reader(open(f"{filepath}",'r'))) #读取的csv以每行为一个列表的形式储存在一个大列表

print("Loading the file, wait......")
file0 = openpyxl.Workbook("Temporary empty file.xlsx")
all_event_sheet0 = file0.create_sheet("all event")
paired_event_sheet0 = file0.create_sheet("paired event")
file0.save("Temporary empty file.xlsx")  # 先创建再读取，否则无法写入
file = openpyxl.load_workbook("Temporary empty file.xlsx")
sheet = file["all event"]
event_sheet = file["paired event"]
root.destroy()

print("Data processing......")
data_column_list = [[],[]]
data_column_list[0].append("107Ag");data_column_list[1].append("109Ag")
for row in data_row_list[1:]:
    for column in range(0, 2):
        data_column_list[column].append(row[column])
'''输入dwell和settling time参数'''


tk = tkinter.Tk()
tk.attributes("-topmost", True)
tk.title('Time resolution')
tk.geometry('500x300') # 设置窗口大小
tk.resizable(0, 0) # 设置窗口宽高固定

tkinter.Label(tk, text='Dwell time (μs)', font=('Arial', 10)).place(x=120, y=50, anchor='w')
tkinter.Label(tk, text='Settling time (μs)', font=('Arial', 10)).place(x=120, y=100, anchor='w')

var_1 = tkinter.StringVar()
var_1.set("select")
tkinter.OptionMenu(tk, var_1, '10', '50', '100', '150', '200').place(x=270, y=50, anchor='w')
var_2 = tkinter.StringVar()
var_2.set("select")
tkinter.OptionMenu(tk, var_2, '10', '50', '100', '150', '200').place(x=270, y=100, anchor='w')

on_hit = True
a,b = '',''
def getValue():
    global on_hit,a,b
    if on_hit == False:
        on_hit = True
        a = var_1.get()
        b = var_2.get()
    else:
        on_hit = False
        tk.destroy()
    return a, b
tkinter.Button(tk, text='ok', width=15, pady=5, command=getValue).place(x=350, y=250, anchor='nw')
tkinter.mainloop()
time_resolution = getValue()


dwell_time = int(time_resolution[0])
settling_time = int(time_resolution[1])
sum_points = int(3 // (2*0.001*(dwell_time + settling_time)))

# '''生成新的积分为3ms的响应列表'''
# sum_column_list = []
# for i in range(0,2):
#     column = data_column_list[i]
#     sum_column = []
#     sum_column.append(column[0])  # 表头
#     for j in range(1,len(column)):
#         column[j] = float(column[j])
#     for j in range(1,len(column),sum_points):
#         if j+sum_points > len(column):
#             break
#         sum_intensity = numpy.sum(column[j:j + sum_points])
#         sum_column.append(sum_intensity)
#     sum_column_list.append(sum_column)

window = 1000*sum_points
row_numb = len(data_column_list[0])

event_column_list = []
for isotope in range(0,2):
    isotope_column_list = data_column_list[isotope]
    col_complete = row_numb * [None]  # 创建一个全为None的列表，后面识别出的事件强度填在相应的位置
    for i in range(1, row_numb, window):
        col = []
        for j in range(i, i + window):
            if j > (len(isotope_column_list) - 1):
                if len(isotope_column_list) - i > 500:
                    break
                else:
                    col = [None]
                    break
            cell = isotope_column_list[j]
            if cell != None:
                cell = float(cell)
            col.append(cell)

        if all(cell is None for cell in col):  # 行底端没有数据了，列表全空，意味着这一部分的计算结束了
            break

        cutoff = 100000
        upper_cutoff = cutoff + 1
        while cutoff != upper_cutoff:  # 1000的window内计算阈值
            col_filtered = list(filter(lambda x: x is not None, col))
            mean = numpy.average(col_filtered)
            std = numpy.std(col_filtered)
            upper_cutoff = cutoff
            # cutoff = mean + 3.29 * std + 2.71;threhold_method = 'IUPAC'
            # cutoff = mean + 3 * std;threhold_method = 'μ+3σ'
            cutoff = mean + 5 * std;threhold_method = 'μ+5σ'
            # cutoff = mean + 7 * std;threhold_method = 'μ+7σ'

            for k in range(0, len(col)):  # 迭代筛出事件，放到前面创建的全为None的列表中
                if col[k] != None and col[k] > cutoff:
                    col_complete[i + k] = col[k]
                    col[k] = None

        col_filtered = list(filter(lambda x: x is not None, col))  # 扣去背景值
        bgd = numpy.average(col_filtered)
        for k in range(0, window):
            if (i+k) >= row_numb:
                break
            if col_complete[i + k] != None:
                col_complete[i + k] = col_complete[i + k] - bgd


    event_column_list.append(col_complete)

col107 =copy.deepcopy(event_column_list[0])
col109 =copy.deepcopy(event_column_list[1])
print(col107[:100])
print(col109[:100])

unpaired_col107 = [None]*(row_numb+1)
unpaired_col109 = [None]*(row_numb+1)
paired_col107 = [None]*(row_numb+1)
paired_col109 = [None]*(row_numb+1)
unpaired_list107 = [[], []]
unpaired_list109 = [[], []]
paired_list = [[], [], []]
i=0
while i < row_numb-2:
    i += 1
    if col107[i] != None:
        if col109[i] != None:
            if  col107[i+1] == None:
                paired_col107[i] = col107[i]
                paired_col109[i] = col109[i]

            elif col107[i+1] != None and col109[i+1] == None:
                paired_col107[i] = col107[i]+col107[i+1]
                paired_col109[i] = col109[i]
                col107[i+1] = None

            elif col107[i + 1] != None and col109[i + 1] != None:
                col107[i+1] = col107[i]+col107[i+1]
                col109[i+1] = col109[i]+col109[i+1]

        if col109[i] == None:
            if col107[i+1] == None and col109[i] == None:
                unpaired_col107[i] = col107[i]

    else:
        if col109[i] != None:
            if col107[i+1] == None:
                unpaired_col109[i] = col109[i]
            elif col107[i+1] != None and col109[i + 1] == None:
                paired_col107[i] = col107[i+1]
                paired_col109[i] = col109[i]
                col107[i+1] = None
            elif col107[i+1] != None and col109[i + 1] != None:
                col107[i + 1] = col107[i + 1]
                col109[i + 1] = col109[i] + col109[i + 1]

for i in range(0, row_numb):
    if unpaired_col107[i] != None:
        sheet.cell(row=i + 1, column=5, value=unpaired_col107[i])
        sheet.cell(row=i + 1, column=7, value='Only107')
        unpaired_list107[0].append(unpaired_col107[i])
        unpaired_list107[1].append('Only107')
    elif unpaired_col109[i] != None:
        sheet.cell(row=i + 1, column=6, value=unpaired_col109[i])
        sheet.cell(row=i + 1, column=7, value='Only109')
        unpaired_list109[0].append(unpaired_col109[i])
        unpaired_list109[1].append('Only109')
    elif paired_col107[i] != None:
        sheet.cell(row=i + 1, column=5, value=paired_col107[i])
        sheet.cell(row=i + 1, column=6, value=paired_col109[i])
        ratio = paired_col109[i] / paired_col107[i]
        sheet.cell(row=i + 1, column=7, value=ratio)
        paired_list[0].append(paired_col107[i])
        paired_list[1].append(paired_col109[i])
        paired_list[2].append(ratio)

for i in range(0,len(data_column_list[0])):
    sheet.cell(i+1,1,data_column_list[0][i])
    sheet.cell(i+1,2,data_column_list[1][i])
for i in range(0,len(event_column_list[0])):
    sheet.cell(i+1,3,event_column_list[0][i])
    sheet.cell(i+1,4,event_column_list[1][i])
for i in range(0,len(paired_list[0])):
    event_sheet.cell(i+2,1,paired_list[0][i])
    event_sheet.cell(i+2,2,paired_list[1][i])
    event_sheet.cell(i+2,3,paired_list[2][i])
for i in range(0, len(unpaired_list107[0])):
    event_sheet.cell(i+2,5,unpaired_list107[0][i])
for i in range(0, len(unpaired_list109[0])):
    event_sheet.cell(i+2,6,unpaired_list109[0][i])

sheet.cell(1,3,'event_107Ag')
sheet.cell(1,4,'event_109Ag')
sheet.cell(1,7,'event_info')
event_sheet.cell(1,1,'107Ag')
event_sheet.cell(1,2,'109Ag')
event_sheet.cell(1,3,'109/107Ag')
event_sheet.cell(1,5,'107Only')
event_sheet.cell(1,6,'109Only')

print("The file is saving, wait......")
file.save(f"{file_name}_{threhold_method}.xlsx")
os.remove("Temporary empty file.xlsx")
print("Special for dual-isotope sp/sc-ICP-MS by Perkin Elmer")
